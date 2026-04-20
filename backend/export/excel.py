from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session
from models import Record


def export_records_to_excel(db: Session) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Records"

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers: 日期时间, 类别, 金额, 描述, AI置信度, 状态, 来源
    headers = ["日期时间", "类别", "金额", "描述", "AI置信度", "状态", "来源"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    records = db.query(Record).order_by(Record.created_at.desc()).all()
    for row_idx, record in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=record.created_at)
        ws.cell(row=row_idx, column=2, value=record.category)
        ws.cell(row=row_idx, column=3, value=record.amount)
        ws.cell(row=row_idx, column=4, value=record.description or "")
        ws.cell(row=row_idx, column=5, value=record.ai_confidence)
        ws.cell(row=row_idx, column=6, value=record.status)
        ws.cell(row=row_idx, column=7, value=record.source)

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 10

    return wb
